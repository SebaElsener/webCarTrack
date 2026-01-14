#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

output_path = Path("./Plantilla_Reporte_Pro.xlsx")
output_path.parent.mkdir(exist_ok=True)

wb = Workbook()

# ----------------------------
# Hoja Datos
# ----------------------------
ws_data = wb.active
ws_data.title = "Datos"

cols = ["Fecha", "Marca", "Modelo", "VIN", "Área", "Avería"]
ws_data.append(cols)

# Crear tabla para pivot
tab = Table(displayName="tblDatos", ref=f"A1:F1")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws_data.add_table(tab)

# ----------------------------
# Hoja Top Areas
# ----------------------------
ws_top_areas = wb.create_sheet("Top Areas")
ws_top_areas.append(["Área", "Casos"])

# Pivot manual (openpyxl no genera pivot real, se deja como plantilla para Excel)
# Excel refrescará automáticamente al abrir

chart1 = BarChart()
chart1.title = "Top 5 Áreas dañadas"
chart1.y_axis.title = "Casos"
chart1.x_axis.title = "Área"
chart1.height = 10
chart1.width = 20
ws_top_areas.add_chart(chart1, "D2")

# ----------------------------
# Hoja Top Averías
# ----------------------------
ws_top_av = wb.create_sheet("Top Averías")
ws_top_av.append(["Avería", "Casos"])

chart2 = BarChart()
chart2.title = "Top 5 Tipos de daño"
chart2.y_axis.title = "Casos"
chart2.x_axis.title = "Avería"
chart2.height = 10
chart2.width = 20
ws_top_av.add_chart(chart2, "D2")

# ----------------------------
# Hoja Evolución
# ----------------------------
ws_evo = wb.create_sheet("Evolución por Fecha")
ws_evo.append(["Fecha", "Casos"])

chart3 = LineChart()
chart3.title = "Evolución de daños por fecha"
chart3.y_axis.title = "Casos"
chart3.x_axis.title = "Fecha"
chart3.height = 10
chart3.width = 20
ws_evo.add_chart(chart3, "D2")

# ----------------------------
# Guardar plantilla
# ----------------------------
wb.save(output_path)
print(f"Plantilla PRO generada en: {output_path}")
