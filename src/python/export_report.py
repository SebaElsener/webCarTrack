#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, Reference, LineChart
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill

HEADERS_DATOS = [
    "Fecha", "Marca", "Modelo", "VIN",
    "Area", "Avería", "Gravedad",
    "Observación", "Batea", "Clima",
    "Movimiento", "Usuario"
]

# =====================================================
# UTILIDADES
# =====================================================
def resaltar_header(ws, row=3, start_col=1, end_col=12):
    thick = Side(style="medium")
    border = Border(bottom=thick)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.border = border

def ajustar_columnas_clave(ws):
    # Anchos óptimos por columna (A=1, B=2, ...)
    widths = {
        1: 9,  # Fecha
        2: 7,  # Marca
        3: 8,  # Modelo
        4: 15,  # VIN
        5: 17,  # Area
        6: 11,  # Avería
        7: 7,  # Gravedad
        8: 25,  # Observación
        9: 6,  # Batea
        10: 6,  # Clima
        11: 6,  # Movimiento
        12: 6,  # User
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Alinear y ajustar texto solo donde importa
    for row in ws.iter_rows(min_row=3):  # datos (salta título + header)
        # Fecha centrada
        row[0].alignment = Alignment(horizontal="center", vertical="center")

        # Marca / Modelo
        row[1].alignment = Alignment(vertical="top")
        row[2].alignment = Alignment(vertical="top")

        # VIN monoespaciado y centrado
        row[3].alignment = Alignment(horizontal="center", vertical="center")

        # Área / Avería con wrap
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")

from openpyxl.styles import Alignment

def ajustar_texto(ws, desde_fila=1):
    for row in ws.iter_rows(min_row=desde_fila):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

def ajustar_altura_filas(ws, min_height=20):
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                max_lines = max(
                    max_lines,
                    str(cell.value).count("\n") + 1
                )
        ws.row_dimensions[row[0].row].height = max(
            min_height * max_lines,
            min_height
        )

def agregar_titulo_hoja(ws, ancho_columnas):
    end_col = get_column_letter(ancho_columnas)
    ws.merge_cells(f"A1:{end_col}1")

    cell = ws["A1"]
    cell.value = ws.title
    cell.font = Font(name=FONT_BASE, bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 8  # ⭐ fila separadora

    ws.print_title_rows = "1:1"
    ws.print_options.horizontalCentered = True


def auto_ajustar_columnas(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(
            max(max_length + 2, min_width),
            max_width
        )

def configurar_impresion(ws):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.horizontalCentered = True

    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


# =====================================================
# ARGUMENTOS
# =====================================================

if len(sys.argv) < 3:
    print("Uso: python export_report.py <output_path_base> <json_datos>")
    sys.exit(1)

output_base = Path(sys.argv[1])
payload = json.loads(sys.argv[2])
output_base.parent.mkdir(parents=True, exist_ok=True)

datos_tabla = payload.get("datos", [])
top_areas = payload.get("topAreas", [])
top_averias = payload.get("topAverias", [])
evolucion = payload.get("evolucion", [])

wb = Workbook()

# =====================================================
# HOJA DATOS
# =====================================================

FONT_BASE = "Calibri"
ws_data = wb.active
ws_data.title = "Datos"
headers = HEADERS_DATOS

agregar_titulo_hoja(ws_data, ancho_columnas=len(headers))

ws_data.append([])            # fila 2 vacía
ws_data.append(headers)       # fila 3

header_fill = PatternFill("solid", fgColor="E9ECEF")
ws_data.row_dimensions[3].height = 26

for cell in ws_data[3]:
    cell.fill = header_fill

for d in datos_tabla:
    fecha = d.get("fecha")
    try:
        fecha = datetime.fromisoformat(fecha.replace("Z", "")).date()
    except:
        pass

    # lista de daños (puede venir vacía)
    damages = d.get("damages") or []

# VIN sin daños (opcional)
    areas = (d.get("areas") or "").split(", ")
    averias = (d.get("averias") or "").split(", ")
    gravedades = (d.get("gravedades") or "").split(", ")
    observaciones = [
    o.strip()
    for o in (d.get("obs") or "").split("|")
    if o.strip()
    ]

    max_len = max(
        len(areas),
        len(averias),
        len(gravedades),
        1
    )

    for i in range(max_len):
        ws_data.append([
            fecha if i == 0 else "",          # solo en la primera fila
            d.get("marca", "") if i == 0 else "",
            d.get("modelo", "") if i == 0 else "",
            d.get("vin", ""),
            areas[i] if i < len(areas) else "",
            averias[i] if i < len(averias) else "",
            gravedades[i] if i < len(gravedades) else "",
            observaciones[i] if i < len(observaciones) else "",
            d.get("batea", "") if i == 0 else "",
            d.get("clima", "") if i == 0 else "",
            d.get("movimiento", "") if i == 0 else "",
            d.get("user", "") if i == 0 else "",
        ])

last_col_letter = get_column_letter(len(headers))
start_row = 3
end_row = ws_data.max_row

tab = Table(
    displayName="TablaDatos",
    ref=f"A{start_row}:{last_col_letter}{end_row}"
)

tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)

ws_data.add_table(tab)
ws_data.freeze_panes = "A4"

# Header repetido en cada página del PDF
ws_data.print_title_rows = "3:3"

# 🔽 Tipografía compacta para PDF
font_header = Font(name=FONT_BASE, size=9, bold=True)
font_data = Font(name=FONT_BASE, size=8)

for row in ws_data.iter_rows(min_row=3):  # header + datos
    for cell in row:
        if cell.row == 3:
            cell.font = font_header
        else:
            cell.font = font_data

fill_even = PatternFill("solid", fgColor="F4F6F8")  # gris claro tipo UI
fill_odd = PatternFill("solid", fgColor="FFFFFF")

current_vin = None
vin_index = -1  # contador de VIN distintos

for row in ws_data.iter_rows(min_row=4, max_row=ws_data.max_row):
    vin_cell = row[3]  # columna VIN (D)

    if vin_cell.value != current_vin:
        current_vin = vin_cell.value
        vin_index += 1

    fill = fill_even if vin_index % 2 == 0 else fill_odd

    for cell in row:
        cell.fill = fill

# 🔹 Merge visual del VIN (sin merge_cells)
current_vin = None
first_row_of_vin = None

for row_idx in range(4, ws_data.max_row + 1):
    vin_cell = ws_data.cell(row=row_idx, column=4)  # VIN = columna D

    if vin_cell.value != current_vin:
        # nuevo VIN
        current_vin = vin_cell.value
        first_row_of_vin = row_idx

        # centrar verticalmente el VIN
        vin_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
    else:
        # mismo VIN → ocultar valor
        vin_cell.value = ""

ajustar_columnas_clave(ws_data)
ajustar_texto(ws_data, desde_fila=3)
ajustar_altura_filas(ws_data)
configurar_impresion(ws_data)

# filas reales de datos (incluye header)
first_row = 2   # header
last_row = ws_data.max_row
first_col = 1
last_col = 6

resaltar_header(ws_data, row=3, end_col=len(headers))

# =====================================================
# TOP ÁREAS
# =====================================================

ws_areas = wb.create_sheet("Top Areas")
agregar_titulo_hoja(ws_areas, ancho_columnas=2)

ws_areas.append([])
ws_areas.append(["Area", "Casos"])

for cell in ws_areas[3]:
    cell.font = Font(name=FONT_BASE, size=9, bold=True)

for row in ws_areas.iter_rows(min_row=4):
    for cell in row:
        cell.font = Font(name=FONT_BASE, size=8)

for t in top_areas:
    ws_areas.append([t["label"], t["value"]])

auto_ajustar_columnas(ws_areas)
configurar_impresion(ws_areas)


ws_areas_chart = wb.create_sheet("Top Areas - Gráfico")
agregar_titulo_hoja(ws_areas_chart, ancho_columnas=6)

chart = PieChart()
chart.title = None
chart.style = 10  # look moderno
chart.height = 15
chart.width = 25

labels = Reference(
    ws_areas,
    min_col=1,
    min_row=4,
    max_row=ws_areas.max_row
)

data = Reference(
    ws_areas,
    min_col=2,
    min_row=3,
    max_row=ws_areas.max_row
)

chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

chart.dataLabels = DataLabelList()
chart.dataLabels.showCatName = True
chart.dataLabels.showPercent = True
# chart.dataLabels.showLeaderLines = True
chart.dataLabels.showVal = False


# solo crear gráfico si hay datos reales
if ws_areas.max_row > 3:   # header + al menos 1 fila de datos
    ws_areas_chart.add_chart(chart, "A3")

configurar_impresion(ws_areas_chart)


# =====================================================
# TOP AVERÍAS
# =====================================================

ws_averias = wb.create_sheet("Top Averías")
agregar_titulo_hoja(ws_averias, ancho_columnas=2)

ws_averias.append([])
ws_averias.append(["Avería", "Casos"])

for cell in ws_averias[3]:
    cell.font = Font(name=FONT_BASE, size=9, bold=True)

for row in ws_averias.iter_rows(min_row=4):
    for cell in row:
        cell.font = Font(name=FONT_BASE, size=8)

for t in top_averias:
    ws_averias.append([t["label"], t["value"]])

auto_ajustar_columnas(ws_averias)
configurar_impresion(ws_averias)


ws_averias_chart = wb.create_sheet("Top Averías - Gráfico")
agregar_titulo_hoja(ws_averias_chart, ancho_columnas=6)

chart2 = PieChart()
chart2.title = None
chart2.style = 10
chart2.height = 15
chart2.width = 25

labels2 = Reference(
    ws_averias,
    min_col=1,
    min_row=4,
    max_row=ws_averias.max_row
)

data2 = Reference(
    ws_averias,
    min_col=2,
    min_row=3,
    max_row=ws_averias.max_row
)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(labels2)

chart2.dataLabels = DataLabelList()
chart2.dataLabels.showCatName = True
chart2.dataLabels.showPercent = True
# chart2.dataLabels.showLeaderLines = True
chart2.dataLabels.showVal = False

# solo crear gráfico si hay datos reales
if ws_averias.max_row > 3:   # header + al menos 1 fila de datos
    ws_averias_chart.add_chart(chart2, "A3")

configurar_impresion(ws_averias_chart)


# =====================================================
# EVOLUCIÓN
# =====================================================

ws_evo = wb.create_sheet("Evolución")
agregar_titulo_hoja(ws_evo, ancho_columnas=2)

ws_evo.append([])
ws_evo.append(["Fecha", "Casos"])

for cell in ws_evo[3]:
    cell.font = Font(name=FONT_BASE, size=9, bold=True)

for row in ws_evo.iter_rows(min_row=4):
    for cell in row:
        cell.font = Font(name=FONT_BASE, size=8)

for t in evolucion:
    ws_evo.append([t["fecha"], t["value"]])

auto_ajustar_columnas(ws_evo)
configurar_impresion(ws_evo)


ws_evo_chart = wb.create_sheet("Evolución - Gráfico")
agregar_titulo_hoja(ws_evo_chart, ancho_columnas=12)

chart3 = LineChart()
chart3.title = "Evolución de daños por fecha"
chart3.y_axis.title = "Casos"
chart3.x_axis.title = "Fecha"

data3 = Reference(ws_evo, min_col=2, min_row=4, max_row=ws_evo.max_row)
cats3 = Reference(ws_evo, min_col=1, min_row=4, max_row=ws_evo.max_row)

chart3.add_data(data3, titles_from_data=False)
chart3.set_categories(cats3)
chart3.width = 22
chart3.height = 12

ws_evo_chart.add_chart(chart3, "A3")
configurar_impresion(ws_evo_chart)


# =====================================================
# GUARDAR
# =====================================================

excel_path = output_base
wb.save(excel_path)

print(f"Excel generado correctamente en: {excel_path}")
